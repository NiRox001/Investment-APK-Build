# -*- coding: utf-8 -*-
# Investment Assistant V23.15 - Fix Settings Scroll & Layout
# 修复1：增加设置页卡片的高度，解决内容溢出导致的“下拉回弹”和“顶部显示不全”的问题
# 修复2：在 Security 标题上方增加间距，使其垂直居中/下移
# 保持：V23.14 的所有功能 (开屏提示、新图标、实时刷新等)

from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.scrollview import ScrollView
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.uix.textinput import TextInput
from kivy.uix.popup import Popup
from kivy.uix.screenmanager import ScreenManager, Screen
from kivy.core.window import Window
from kivy.clock import Clock
from kivy.metrics import dp, sp
from kivy.graphics import Color, RoundedRectangle, Line, Ellipse, Rectangle, Mesh, Triangle
from kivy.properties import ListProperty, NumericProperty, ObjectProperty, StringProperty
from kivy.animation import Animation
from kivy.uix.behaviors import ButtonBehavior
from kivy.uix.widget import Widget

import os
import sys
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import pandas as pd
from datetime import datetime
import threading
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import shutil
import json
import hashlib
import math


# === 顶部通知条组件 (Toast) ===
class Toast(Label):
    def __init__(self, text, duration=2.5, bg_color=(0.2, 0.2, 0.2, 0.9), **kwargs):
        super().__init__(**kwargs)
        self.text = text
        self.size_hint = (0.9, None)
        self.height = dp(40)
        self.pos_hint = {'top': 1.2, 'center_x': 0.5}
        self.color = (1, 1, 1, 1)
        self.bold = True
        self.font_size = '14sp'

        with self.canvas.before:
            Color(*bg_color)
            self.bg_rect = RoundedRectangle(pos=self.pos, size=self.size, radius=[dp(20)])

        self.bind(pos=self.update_rect, size=self.update_rect)
        self.show(duration)

    def update_rect(self, *args):
        self.bg_rect.pos = self.pos
        self.bg_rect.size = self.size

    def show(self, duration):
        anim_in = Animation(pos_hint={'top': 0.98}, duration=0.3, t='out_back')
        anim_out = Animation(pos_hint={'top': 1.2}, duration=0.3, t='in_back')
        anim_in.bind(on_complete=lambda *args: Clock.schedule_once(lambda dt: anim_out.start(self), duration))
        anim_out.bind(on_complete=lambda *args: self.parent.remove_widget(self) if self.parent else None)

        app = App.get_running_app()
        if app and app.root:
            app.root.add_widget(self)
            anim_in.start(self)


class GradientButton(Button):
    color1 = ListProperty([0.2, 0.6, 0.86, 1])
    color2 = ListProperty([0.15, 0.45, 0.72, 1])

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.background_normal = ''
        self.background_down = ''
        self.background_color = (0, 0, 0, 0)

        self.bind(pos=self.update_canvas, size=self.update_canvas)
        self.bind(color1=self.update_canvas, color2=self.update_canvas)
        self.update_canvas()

    def update_canvas(self, *args):
        self.canvas.before.clear()
        with self.canvas.before:
            Color(*self.color1)
            RoundedRectangle(pos=self.pos, size=self.size, radius=[dp(12)])

    def on_press(self):
        anim = Animation(opacity=0.7, duration=0.1)
        anim.start(self)

    def on_release(self):
        anim = Animation(opacity=1.0, duration=0.1)
        anim.start(self)


class ModernCard(BoxLayout):
    def __init__(self, bg_color=(1, 1, 1, 1), elevation=2, **kwargs):
        super().__init__(**kwargs)
        self.padding = dp(15)

        with self.canvas.before:
            if elevation > 0:
                Color(0, 0, 0, 0.1)
                self.shadow = RoundedRectangle(
                    pos=(self.x + dp(2), self.y - dp(2)),
                    size=self.size,
                    radius=[dp(15)]
                )

            Color(*bg_color)
            self.rect = RoundedRectangle(pos=self.pos, size=self.size, radius=[dp(15)])

        self.bind(pos=self.update_rect, size=self.update_rect)

    def update_rect(self, *args):
        self.rect.pos = self.pos
        self.rect.size = self.size
        if hasattr(self, 'shadow'):
            self.shadow.pos = (self.x + dp(2), self.y - dp(2))
            self.shadow.size = self.size


class StyledTextInput(BoxLayout):
    def __init__(self, text='', multiline=False, input_filter=None, font_size='14sp', password=False, **kwargs):
        super().__init__(**kwargs)
        self.orientation = 'horizontal'
        self.size_hint_y = None
        self.height = dp(45)

        with self.canvas.before:
            Color(0.95, 0.95, 0.95, 1)
            self.bg_rect = RoundedRectangle(pos=self.pos, size=self.size, radius=[dp(10)])

        self.bind(pos=self.update_bg, size=self.update_bg)

        self.textinput = TextInput(
            text=text,
            multiline=multiline,
            input_filter=input_filter,
            font_size=font_size,
            password=password,
            background_normal='',
            background_active='',
            background_color=(0, 0, 0, 0),
            foreground_color=(0.2, 0.2, 0.2, 1),
            cursor_color=(0.2, 0.6, 0.86, 1),
            selection_color=(0.2, 0.6, 0.86, 0.3),
            padding=[dp(15), dp(16), dp(15), dp(16)]
        )

        self.add_widget(self.textinput)

    def update_bg(self, *args):
        self.bg_rect.pos = self.pos
        self.bg_rect.size = self.size

    @property
    def text(self):
        return self.textinput.text

    @text.setter
    def text(self, value):
        self.textinput.text = value


class ProgressRing(BoxLayout):
    progress = NumericProperty(0)
    color = ListProperty([0.2, 0.6, 0.86, 1])

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.bind(progress=self.update_ring, pos=self.update_ring, size=self.update_ring)
        self.update_ring()

    def update_ring(self, *args):
        self.canvas.before.clear()

        center_x = self.center_x
        center_y = self.center_y
        radius = min(self.width, self.height) / 2 - dp(10)

        with self.canvas.before:
            Color(0.9, 0.9, 0.9, 1)
            Line(circle=(center_x, center_y, radius), width=dp(8))

            Color(*self.color)
            angle = 360 * (self.progress / 100)
            Line(circle=(center_x, center_y, radius, 0, angle), width=dp(8))


class DoubleLineChart(BoxLayout):
    def __init__(self, market_data=None, cost_data=None, **kwargs):
        super().__init__(**kwargs)
        self.market_data = market_data or []
        self.cost_data = cost_data or []
        self.orientation = 'vertical'

        legend = BoxLayout(size_hint_y=None, height=dp(20), spacing=dp(10))
        legend.add_widget(Label(text='', size_hint_x=0.2))

        l1 = BoxLayout(size_hint_x=None, width=dp(80))
        with l1.canvas:
            Color(0.2, 0.6, 0.86, 1)
            Line(points=[l1.x, l1.center_y, l1.right, l1.center_y], width=2)
        l1.bind(pos=lambda i, v: self.draw_chart(), size=lambda i, v: self.draw_chart())
        legend.add_widget(l1)
        legend.add_widget(
            Label(text='Mkt Value', font_size='10sp', color=(0.5, 0.5, 0.5, 1), size_hint_x=None, width=dp(50)))

        l2 = BoxLayout(size_hint_x=None, width=dp(80))
        with l2.canvas:
            Color(0.7, 0.7, 0.7, 1)
            Line(points=[l2.x, l2.center_y, l2.right, l2.center_y], width=1.5)
        legend.add_widget(l2)
        legend.add_widget(
            Label(text='Invested', font_size='10sp', color=(0.5, 0.5, 0.5, 1), size_hint_x=None, width=dp(50)))
        legend.add_widget(Label(text='', size_hint_x=0.2))

        self.add_widget(legend)

        self.chart_area = Widget()
        self.add_widget(self.chart_area)

        self.bind(pos=self.draw_chart, size=self.draw_chart)
        self.draw_chart()

    def set_data(self, market_data, cost_data=None):
        self.market_data = market_data or []
        self.cost_data = cost_data or []
        self.draw_chart()

    def _catmull_rom_spline(self, p0, p1, p2, p3, n_points=20):
        points = []
        t0 = 0
        t1 = 1
        t2 = 2
        t3 = 3

        for i in range(n_points):
            t = t1 + (t2 - t1) * (i / n_points)

            a1 = (t1 - t) / (t1 - t0) * p0 + (t - t0) / (t1 - t0) * p1
            a2 = (t2 - t) / (t2 - t1) * p1 + (t - t1) / (t2 - t1) * p2
            a3 = (t3 - t) / (t3 - t2) * p2 + (t - t2) / (t3 - t2) * p3

            b1 = (t2 - t) / (t2 - t0) * a1 + (t - t0) / (t2 - t0) * a2
            b2 = (t3 - t) / (t3 - t1) * a2 + (t - t1) / (t3 - t1) * a3

            c = (t2 - t) / (t2 - t1) * b1 + (t - t1) / (t2 - t1) * b2
            points.append(c)

        return points

    def _calculate_smooth_path(self, raw_points):
        if len(raw_points) < 2:
            return raw_points

        points = [raw_points[0]] + raw_points + [raw_points[-1]]

        smooth_points = []
        for i in range(len(points) - 3):
            p0, p1, p2, p3 = points[i], points[i + 1], points[i + 2], points[i + 3]

            xs = self._catmull_rom_spline(p0[0], p1[0], p2[0], p3[0])
            ys = self._catmull_rom_spline(p0[1], p1[1], p2[1], p3[1])

            for x, y in zip(xs, ys):
                smooth_points.append(x)
                smooth_points.append(y)

        smooth_points.append(raw_points[-1][0])
        smooth_points.append(raw_points[-1][1])

        return smooth_points

    def draw_chart(self, *args):
        if not self.market_data or len(self.market_data) < 2:
            self.chart_area.canvas.clear()
            return

        self.chart_area.canvas.clear()

        all_values = self.market_data + self.cost_data
        if not all_values: return

        max_val = max(all_values)
        min_val = min(all_values)
        val_range = max_val - min_val if max_val != min_val else 1

        val_range = val_range * 1.15
        min_val = min_val * 0.95

        x_start = self.chart_area.x + dp(10)
        y_start = self.chart_area.y + dp(10)
        width = self.chart_area.width - dp(20)
        height = self.chart_area.height - dp(20)

        def get_raw_points(data):
            pts = []
            if not data: return []
            x_step = width / (len(data) - 1)
            for i, val in enumerate(data):
                px = x_start + i * x_step
                py = y_start + ((val - min_val) / val_range) * height
                pts.append((px, py))
            return pts

        raw_market_pts = get_raw_points(self.market_data)
        raw_cost_pts = get_raw_points(self.cost_data)

        smooth_market_flat = self._calculate_smooth_path(raw_market_pts)
        smooth_cost_flat = self._calculate_smooth_path(raw_cost_pts)

        with self.chart_area.canvas:
            Color(0.92, 0.92, 0.92, 1)
            for i in range(5):
                y = y_start + (height / 4) * i
                Line(points=[x_start, y, x_start + width, y], width=1)

            if smooth_market_flat:
                vertices = []
                indices = []
                pt_count = len(smooth_market_flat) // 2
                for i in range(pt_count):
                    x = smooth_market_flat[i * 2]
                    y = smooth_market_flat[i * 2 + 1]
                    vertices.extend([x, y, 0, 0])
                    vertices.extend([x, y_start, 0, 0])

                for i in range(pt_count - 1):
                    idx = i * 2
                    indices.extend([idx, idx + 1, idx + 2, idx + 1, idx + 2, idx + 3])

                Color(0.2, 0.6, 0.86, 0.15)
                Mesh(vertices=vertices, indices=indices, mode='triangles')

            if smooth_cost_flat:
                Color(0.7, 0.7, 0.7, 1)
                Line(points=smooth_cost_flat, width=1.5)

            if smooth_market_flat:
                Color(0.2, 0.6, 0.86, 1)
                Line(points=smooth_market_flat, width=2.2)

            if raw_market_pts:
                Color(1, 1, 1, 1)
                lx, ly = raw_market_pts[-1]
                Ellipse(pos=(lx - dp(5), ly - dp(5)), size=(dp(10), dp(10)))
                Color(0.2, 0.6, 0.86, 1)
                Ellipse(pos=(lx - dp(3), ly - dp(3)), size=(dp(6), dp(6)))


class SimpleLineChart(BoxLayout):
    def __init__(self, data_points=None, **kwargs):
        super().__init__(**kwargs)
        self.data_points = data_points or []
        self.bind(pos=self.draw_chart, size=self.draw_chart)
        self.draw_chart()

    def set_data(self, data_points):
        self.data_points = data_points
        self.draw_chart()

    def draw_chart(self, *args):
        if not self.data_points or len(self.data_points) < 2:
            return
        self.canvas.before.clear()
        max_val = max(self.data_points)
        min_val = min(self.data_points)
        val_range = max_val - min_val if max_val != min_val else 1
        width = self.width - dp(40)
        height = self.height - dp(40)
        x_step = width / (len(self.data_points) - 1)
        points = []
        for i, val in enumerate(self.data_points):
            x = self.x + dp(20) + i * x_step
            y = self.y + dp(20) + ((val - min_val) / val_range) * height
            points.extend([x, y])
        with self.canvas.before:
            Color(0.9, 0.9, 0.9, 1)
            for i in range(5):
                y = self.y + dp(20) + (height / 4) * i
                Line(points=[self.x + dp(20), y, self.x + width + dp(20), y], width=1)
            Color(0.2, 0.6, 0.86, 1)
            Line(points=points, width=2)
            for i in range(0, len(points), 2):
                Color(0.15, 0.55, 0.82, 1)
                Ellipse(pos=(points[i] - dp(3), points[i + 1] - dp(3)), size=(dp(6), dp(6)))


class InvestmentBackend:
    def __init__(self):
        # --- 修改开始 ---
        # 获取安卓应用的专用读写目录
        app = App.get_running_app()
        if app:
            self.base_dir = app.user_data_dir
        else:
            # 电脑端测试时的回退目录
            self.base_dir = os.path.dirname(os.path.abspath(__file__))
        # --- 修改结束 ---

        self.excel_path = os.path.join(self.base_dir, 'investment_records.xlsx')
        self.backup_dir = os.path.join(self.base_dir, 'backup')
        self.config_path = os.path.join(self.base_dir, 'config.json')
        # ... 后面的代码保持不变

        if not os.path.exists(self.backup_dir):
            os.makedirs(self.backup_dir)

        self.config = {
            'symbol_a': 'VOO',
            'symbol_b': 'QQQM',
            'fire_goal': 1000000.0,
            'bonus_fund_total': 10000.0,
            'bonus_fund_remaining': 10000.0,
            'reset_password_hash': self._hash_password('1234'),
            'auto_refresh_interval': 60  # 默认 60 分钟
        }

        self.latest_prices = None

        self.load_config()
        self.months_count = 0
        self.reload_status()

    def _hash_password(self, password):
        return hashlib.sha256(password.encode()).hexdigest()

    def verify_reset_password(self, password):
        input_hash = self._hash_password(password)
        stored_hash = self.config.get('reset_password_hash', self._hash_password('1234'))
        return input_hash == stored_hash

    def set_reset_password(self, new_password):
        self.config['reset_password_hash'] = self._hash_password(new_password)
        self.save_config()

    def load_config(self):
        if os.path.exists(self.config_path):
            try:
                with open(self.config_path, 'r', encoding='utf-8') as f:
                    loaded = json.load(f)
                    self.config.update(loaded)
            except:
                pass

    def save_config(self):
        try:
            with open(self.config_path, 'w', encoding='utf-8') as f:
                json.dump(self.config, f, indent=2)
        except:
            pass

    def reload_status(self):
        if os.path.exists(self.excel_path):
            try:
                df = pd.read_excel(self.excel_path)
                self.months_count = len(df)
                return True
            except:
                return False
        else:
            self.months_count = 0
            return True

    def get_all_records(self):
        if not os.path.exists(self.excel_path):
            return None
        try:
            df = pd.read_excel(self.excel_path)
            return df
        except:
            return None

    def get_chart_data(self):
        df = self.get_all_records()
        if df is None or len(df) == 0:
            return None

        investment_list = [float(x) for x in df['Total_Investment'].tolist()]
        market_value_list = [float(x) for x in df['Market_Value'].tolist()] if 'Market_Value' in df.columns else []
        profit_list = [float(x) for x in df['Profit'].tolist()] if 'Profit' in df.columns else []

        return {
            'dates': df['Date'].tolist(),
            'investment': investment_list,
            'market_value': market_value_list,
            'profit': profit_list
        }

    def undo_last_transaction(self):
        if not os.path.exists(self.excel_path):
            return False, "No data file"

        try:
            df = pd.read_excel(self.excel_path)
            if len(df) == 0:
                return False, "No records"

            last_bonus = df.iloc[-1].get('Bonus', 0)
            if pd.notnull(last_bonus) and last_bonus > 0:
                self.config['bonus_fund_remaining'] += float(last_bonus)
                self.save_config()

            df = df.iloc[:-1]
            df.to_excel(self.excel_path, index=False)
            self._format_excel()
            self.reload_status()
            return True, f"Success! Refunded ${float(last_bonus):.2f}"
        except Exception as e:
            return False, f"Failed: {str(e)}"

    def _direct_request(self, url, params=None):
        session = requests.Session()
        session.trust_env = False
        retry = Retry(total=3, backoff_factor=0.5, status_forcelist=[500, 502, 503, 504])
        adapter = HTTPAdapter(max_retries=retry)
        session.mount('http://', adapter)
        session.mount('https://', adapter)

        headers = {
            'User-Agent': 'Mozilla/5.0 (Linux; Android 10) AppleWebKit/537.36',
            'Referer': 'https://finance.sina.com.cn/'
        }

        try:
            res = session.get(url, params=params, headers=headers, timeout=30)
            res.raise_for_status()
            return res
        except:
            return None

    def calculate_valuation_percentile(self, current_price, historical_closes):
        if not historical_closes or len(historical_closes) < 2:
            return 50.0

        count_below = sum(1 for price in historical_closes if price < current_price)
        percentile = (count_below / len(historical_closes)) * 100

        return percentile

    def get_valuation_level(self, percentile):
        if percentile < 10:
            return "Very Low", (0.13, 0.7, 0.38, 1)
        elif percentile < 25:
            return "Low", (0.4, 0.8, 0.4, 1)
        elif percentile < 40:
            return "Fair-Low", (0.2, 0.6, 0.86, 1)
        elif percentile < 60:
            return "Fair", (0.5, 0.5, 0.5, 1)
        elif percentile < 75:
            return "Fair-High", (0.95, 0.61, 0.07, 1)
        elif percentile < 90:
            return "High", (0.91, 0.49, 0.13, 1)
        else:
            return "Very High", (0.91, 0.3, 0.24, 1)

    def calculate_rsi(self, closes, period=14):
        if len(closes) < period + 1:
            return 50

        deltas = [closes[i] - closes[i - 1] for i in range(1, len(closes))]
        gains = [d if d > 0 else 0 for d in deltas]
        losses = [-d if d < 0 else 0 for d in deltas]

        avg_gain = sum(gains[-period:]) / period
        avg_loss = sum(losses[-period:]) / period

        if avg_loss == 0:
            return 100

        rs = avg_gain / avg_loss
        rsi = 100 - (100 / (1 + rs))
        return rsi

    def calculate_macd(self, closes):
        if len(closes) < 26:
            return False

        def ema(data, period):
            multiplier = 2 / (period + 1)
            ema_values = [sum(data[:period]) / period]
            for price in data[period:]:
                ema_values.append((price - ema_values[-1]) * multiplier + ema_values[-1])
            return ema_values[-1]

        ema12 = ema(closes, 12)
        ema26 = ema(closes, 26)
        macd_line = ema12 - ema26

        if len(closes) >= 27:
            prev_ema12 = ema(closes[:-1], 12)
            prev_ema26 = ema(closes[:-1], 26)
            prev_macd = prev_ema12 - prev_ema26
            if macd_line > 0 and macd_line > prev_macd:
                return True

        return False

    def calculate_bollinger_bands(self, closes, period=20, std_dev=2):
        if len(closes) < period:
            return False

        recent_closes = closes[-period:]
        sma = sum(recent_closes) / period
        variance = sum((x - sma) ** 2 for x in recent_closes) / period
        std = variance ** 0.5
        lower_band = sma - (std * std_dev)
        current_price = closes[-1]

        if len(closes) >= period + 1:
            prev_price = closes[-2]
            if prev_price < lower_band and current_price >= lower_band * 0.98:
                return True

        return False

    def get_market_data(self):
        ndx = None
        prices = {}
        error_msg = ""

        try:
            url_ndx = "https://push2his.eastmoney.com/api/qt/stock/kline/get"

            params_monthly = {
                'secid': '100.NDX',
                'fields1': 'f1,f2,f3',
                'fields2': 'f51,f52,f53',
                'klt': '103',
                'fqt': '1',
                'beg': '0',
                'end': '20500000',
                'lmt': '2'
            }
            res_monthly = self._direct_request(url_ndx, params_monthly)

            params_daily = {
                'secid': '100.NDX',
                'fields1': 'f1,f2,f3',
                'fields2': 'f51,f52,f53',
                'klt': '101',
                'fqt': '1',
                'beg': '0',
                'end': '20500000',
                'lmt': '150'
            }
            res_daily = self._direct_request(url_ndx, params_daily)

            params_valuation = {
                'secid': '100.NDX',
                'fields1': 'f1,f2,f3',
                'fields2': 'f51,f52,f53',
                'klt': '101',
                'fqt': '1',
                'beg': '20150101',
                'end': '20500000',
                'lmt': '5000'
            }
            res_valuation = self._direct_request(url_ndx, params_valuation)

            if res_monthly and res_daily:
                data_m = res_monthly.json()
                data_d = res_daily.json()

                if data_m and data_m.get('data') and data_m['data'].get('klines'):
                    klines_m = data_m['data']['klines']
                    curr = float(klines_m[-1].split(',')[2])
                    prev = float(klines_m[-2].split(',')[2]) if len(klines_m) > 1 else curr
                    change_pct = ((curr - prev) / prev) * 100 if prev > 0 else 0

                    ma120 = 0
                    rsi = 50
                    macd_golden = False
                    bb_breakthrough = False

                    if data_d and data_d.get('data') and data_d['data'].get('klines'):
                        klines_d = data_d['data']['klines']
                        closes = [float(k.split(',')[2]) for k in klines_d]

                        if len(closes) >= 120:
                            ma120 = sum(closes[-120:]) / 120
                        elif len(closes) > 0:
                            ma120 = sum(closes) / len(closes)

                        rsi = self.calculate_rsi(closes)
                        macd_golden = self.calculate_macd(closes)
                        bb_breakthrough = self.calculate_bollinger_bands(closes)

                valuation_percentile = 50.0
                valuation_history_closes = []

                if res_valuation:
                    data_v = res_valuation.json()
                    if data_v and data_v.get('data') and data_v['data'].get('klines'):
                        klines_v = data_v['data']['klines']
                        valuation_history_closes = [float(k.split(',')[2]) for k in klines_v]
                        valuation_percentile = self.calculate_valuation_percentile(curr, valuation_history_closes)

                    ndx = {
                        'date': datetime.now().strftime('%Y-%m-%d'),
                        'close': curr,
                        'prev_close': prev,
                        'change': change_pct,
                        'ma120': ma120,
                        'rsi': rsi,
                        'macd_golden': macd_golden,
                        'bb_breakthrough': bb_breakthrough,
                        'valuation_percentile': valuation_percentile,
                        'valuation_history_count': len(valuation_history_closes)
                    }

        except Exception as e:
            error_msg += f"Error: {e}\n"

        if ndx is None:
            ndx = {
                'date': datetime.now().strftime('%Y-%m-%d'),
                'close': 23474.35,
                'prev_close': 23364.0,
                'change': 0.47,
                'ma120': 22365.04,
                'rsi': 48.4,
                'macd_golden': False,
                'bb_breakthrough': False,
                'valuation_percentile': 50.0,
                'valuation_history_count': 0
            }
            error_msg += "Using mock data\n"

        for sym in [self.config['symbol_a'], self.config['symbol_b']]:
            price_found = False

            try:
                sina_sym = f"gb_{sym.lower()}"
                res_sina = self._direct_request(f"http://hq.sinajs.cn/list={sina_sym}")
                if res_sina and '="' in res_sina.text:
                    prices[sym] = float(res_sina.text.split('="')[1].split(',')[1])
                    price_found = True
            except:
                pass

            if not price_found:
                for mkt in ['105', '106']:
                    try:
                        res_em = self._direct_request(
                            "https://push2.eastmoney.com/api/qt/stock/get",
                            {'secid': f'{mkt}.{sym.upper()}', 'fields': 'f43'}
                        )
                        if res_em:
                            d = res_em.json().get('data')
                            if d and d.get('f43') and d['f43'] != '-':
                                val = float(d['f43'])
                                if 10000 < val < 500000:
                                    val = val / 1000
                                prices[sym] = val
                                price_found = True
                                break
                    except:
                        continue

            if not price_found:
                default_prices = {'VOO': 632.60, 'QQQM': 240.0}
                prices[sym] = default_prices.get(sym, 100.0)

        return ndx, prices, error_msg

    def calculate_strategy(self, market_data):
        change_pct = market_data['change']
        current_price = market_data['close']
        ma120 = market_data['ma120']
        rsi = market_data['rsi']
        macd_golden = market_data['macd_golden']
        bb_breakthrough = market_data['bb_breakthrough']

        if change_pct < -15:
            regular_amount = 1000
        elif change_pct < -10:
            regular_amount = 700
        elif change_pct < -5:
            regular_amount = 400
        else:
            regular_amount = 250

        indicator_count = 0
        indicator_details = []
        bonus_amount = 0
        remaining_fund = self.config.get('bonus_fund_remaining', 0)

        if rsi < 30:
            indicator_count += 1
            indicator_details.append("RSI<30")

        if change_pct < -10:
            indicator_count += 1
            indicator_details.append("Drop>10%")

        below_ma120 = (ma120 > 0 and current_price < ma120)

        if below_ma120:
            indicator_count += 1
            indicator_details.append("MA120")

            if macd_golden:
                indicator_count += 1
                indicator_details.append("MACD")

            if bb_breakthrough:
                indicator_count += 1
                indicator_details.append("BB")

        if indicator_count > 0 and remaining_fund > 0:
            trigger_coef_map = {
                1: 0.10,
                2: 0.15,
                3: 0.25,
                4: 0.35,
                5: 0.50
            }
            trigger_coef = trigger_coef_map.get(indicator_count, 0.10)

            if remaining_fund > 8000:
                risk_coef = 1.0
            elif remaining_fund > 5000:
                risk_coef = 0.8
            elif remaining_fund > 2000:
                risk_coef = 0.6
            else:
                risk_coef = 0.4

            calculated = remaining_fund * trigger_coef * risk_coef
            bonus_amount = min(calculated, 2000.0)

        return {
            'regular_amount': regular_amount,
            'bonus_amount': bonus_amount,
            'total_amount': regular_amount + bonus_amount,
            'indicator_count': indicator_count,
            'indicator_details': indicator_details,
            'below_ma120': below_ma120,
            'remaining_fund': remaining_fund,
            'change_pct': change_pct
        }

    def save_transaction(self, market, strategy, details):
        cost_a = details['price_a'] * details['share_a']
        cost_b = details['price_b'] * details['share_b']
        total_cost = cost_a + cost_b

        prev_principal = 0
        if os.path.exists(self.excel_path):
            try:
                df_prev = pd.read_excel(self.excel_path)
                if len(df_prev) > 0:
                    prev_principal = float(df_prev['Total_Investment'].iloc[-1])
            except:
                pass

        curr_principal = prev_principal + total_cost

        record = {
            'Date': market['date'],
            'NDX_Price': market['close'],
            'NDX_Change': f"{market['change']:.2f}%",
            'MA120': market['ma120'],
            'RSI': f"{market['rsi']:.1f}",
            'Valuation': f"{market.get('valuation_percentile', 50):.1f}%",
            'Indicators': strategy['indicator_count'],
            'Details': '+'.join(strategy['indicator_details']) if strategy['indicator_details'] else 'None',
            'Regular': strategy['regular_amount'],
            'Bonus': strategy['bonus_amount'],
            'Monthly_Cost': total_cost,
            'VOO_Price': details['price_a'],
            'VOO_Shares': details['share_a'],
            'QQQM_Price': details['price_b'],
            'QQQM_Shares': details['share_b'],
            'Total_Investment': curr_principal,
            'Bonus_Balance': self.config['bonus_fund_remaining'] - strategy['bonus_amount']
        }

        total_voo = details['share_a']
        total_qqqm = details['share_b']

        if os.path.exists(self.excel_path):
            try:
                hist_df = pd.read_excel(self.excel_path)
                for _, r in hist_df.iterrows():
                    total_voo += float(r.get('VOO_Shares', 0))
                    total_qqqm += float(r.get('QQQM_Shares', 0))
            except:
                pass

        mkt_val = (total_voo * details['price_a']) + (total_qqqm * details['price_b'])
        record['Market_Value'] = mkt_val
        record['Profit'] = mkt_val - curr_principal
        record['Return_Pct'] = (mkt_val - curr_principal) / curr_principal if curr_principal > 0 else 0

        df_new = pd.DataFrame([record])

        if os.path.exists(self.excel_path):
            try:
                df = pd.concat([pd.read_excel(self.excel_path), df_new], ignore_index=True)
            except:
                df = df_new
        else:
            df = df_new

        with pd.ExcelWriter(self.excel_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)

        self.config['bonus_fund_remaining'] -= strategy['bonus_amount']
        self.save_config()
        self.months_count += 1
        self._format_excel()
        self.backup_data()

        return record

    def _format_excel(self):
        try:
            wb = load_workbook(self.excel_path)
            ws = wb.active

            fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            white_font = Font(bold=True, color="FFFFFF")
            border = Border(
                left=Side(style='thin', color="BFBFBF"),
                right=Side(style='thin', color="BFBFBF"),
                top=Side(style='thin', color="BFBFBF"),
                bottom=Side(style='thin', color="BFBFBF")
            )

            for cell in ws[1]:
                cell.fill = fill
                cell.font = white_font

            for row in ws.iter_rows():
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center')

            for i in range(1, ws.max_column + 1):
                ws.column_dimensions[ws.cell(1, i).column_letter].width = 14

            wb.save(self.excel_path)
        except:
            pass

    def backup_data(self):
        if os.path.exists(self.excel_path):
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = os.path.join(self.backup_dir, f"backup_{ts}.xlsx")
            shutil.copy2(self.excel_path, backup_path)
            return True
        return False

    def reset_data(self):
        if os.path.exists(self.excel_path):
            os.remove(self.excel_path)
        self.config['bonus_fund_remaining'] = self.config['bonus_fund_total']
        self.save_config()
        self.reload_status()
        return True

    def get_avg_cost(self, current_prices=None):
        """
        获取持仓成本和市值。
        :param current_prices: (Optional) 包含最新价格的字典 {'VOO': 100, 'QQQM': 200}
                               如果提供，将使用最新价格计算实时市值和收益。
        """
        if not os.path.exists(self.excel_path):
            return None

        try:
            df = pd.read_excel(self.excel_path)
            if len(df) == 0:
                return None

            sym_a = self.config['symbol_a']
            sym_b = self.config['symbol_b']

            # 计算总持仓和总投入
            share_a = df['VOO_Shares'].sum()
            share_b = df['QQQM_Shares'].sum()

            # 计算平均成本
            cost_a_total = df['VOO_Price'].mul(df['VOO_Shares']).sum()
            cost_b_total = df['QQQM_Price'].mul(df['QQQM_Shares']).sum()
            avg_a = cost_a_total / share_a if share_a > 0 else 0
            avg_b = cost_b_total / share_b if share_b > 0 else 0

            # 本金一般取最后一条记录的 Total_Investment
            principal = float(df['Total_Investment'].iloc[-1]) if len(df) > 0 else 0

            # 决定使用历史市值还是实时市值
            if current_prices:
                pa = current_prices.get(sym_a, 0)
                pb = current_prices.get(sym_b, 0)

                # 如果获取到了有效价格
                if pa > 0 or pb > 0:
                    market_val = (share_a * pa) + (share_b * pb)
                    profit = market_val - principal
                else:
                    # 获取价格失败，回退到历史数据
                    market_val = float(df['Market_Value'].iloc[-1]) if 'Market_Value' in df.columns else 0
                    profit = float(df['Profit'].iloc[-1]) if 'Profit' in df.columns else 0
            else:
                # 使用 Excel 中记录的最后一次市值
                market_val = float(df['Market_Value'].iloc[-1]) if 'Market_Value' in df.columns else 0
                profit = float(df['Profit'].iloc[-1]) if 'Profit' in df.columns else 0

            return {
                sym_a: (float(avg_a), float(share_a)),
                sym_b: (float(avg_b), float(share_b)),
                'principal': principal,
                'market_value': market_val,
                'profit': profit
            }
        except:
            return None


# === V23.10 新增：首页统计卡片的绘图图标组件 (StatIcon) ===
class StatIcon(Widget):
    def __init__(self, icon_type, color=(1, 1, 1, 1), **kwargs):
        super().__init__(**kwargs)
        self.icon_type = icon_type
        self.color = color
        self.bind(pos=self.update_canvas, size=self.update_canvas)

    def update_canvas(self, *args):
        self.canvas.clear()
        cx, cy = self.center_x, self.center_y
        w, h = self.width, self.height

        # 调整图标大小
        size = min(w, h) * 0.7

        with self.canvas:
            Color(*self.color)

            if self.icon_type == 'moneybag':  # Invested - 方案 A (钱袋)
                # 钱袋底部 (圆形)
                Ellipse(pos=(cx - size / 2, cy - size / 2), size=(size, size))
                # 钱袋束口 (三角形)
                Triangle(points=[
                    cx - size / 3, cy + size / 3,
                    cx + size / 3, cy + size / 3,
                    cx, cy + size / 1.5
                ])
                # 美元符号 $
                Color(1, 1, 1, 0.5)  # 半透明白色
                # S形曲线模拟
                Line(bezier=[cx + size / 4, cy + size / 4, cx - size / 2, cy, cx + size / 2, cy, cx - size / 4,
                             cy - size / 4], width=1.5)
                # 竖线
                Line(points=[cx, cy + size / 3, cx, cy - size / 3], width=1.5)

            elif self.icon_type == 'trend':  # Market - 趋势线 (保持不变)
                # 坐标轴
                Line(points=[cx - size / 2, cy + size / 2, cx - size / 2, cy - size / 2, cx + size / 2, cy - size / 2],
                     width=dp(1.5))
                # 趋势线 (闪电形状)
                pts = [
                    cx - size / 2, cy - size / 2,
                    cx - size / 6, cy,
                    cx + size / 6, cy - size / 4,
                    cx + size / 2, cy + size / 2
                ]
                Line(points=pts, width=dp(2))
                # 箭头头部
                Line(points=[cx + size / 2 - dp(5), cy + size / 2, cx + size / 2, cy + size / 2, cx + size / 2,
                             cy + size / 2 - dp(5)], width=dp(2))

            elif self.icon_type == 'rocket':  # Profit - 方案 F (火箭)
                # 火箭主体 (椭圆)
                Ellipse(pos=(cx - size / 4, cy - size / 4), size=(size / 2, size), angle=0)
                # 尾翼 (左右三角形)
                Triangle(points=[
                    cx - size / 4, cy,
                    cx - size / 2, cy - size / 2,
                    cx - size / 4, cy - size / 3
                ])
                Triangle(points=[
                    cx + size / 4, cy,
                    cx + size / 2, cy - size / 2,
                    cx + size / 4, cy - size / 3
                ])
                # 舷窗 (白色圆点)
                Color(1, 1, 1, 0.6)
                Ellipse(pos=(cx - size / 8, cy + size / 4), size=(size / 4, size / 4))


# ==================== Dashboard Screen ====================  
class DashboardScreen(Screen):
    def __init__(self, backend, **kwargs):
        super().__init__(**kwargs)
        self.backend = backend
        self.build_ui()

    def build_ui(self):
        layout = BoxLayout(orientation='vertical', spacing=0)

        scroll = ScrollView()
        content = BoxLayout(orientation='vertical', size_hint_y=None, spacing=dp(15), padding=dp(15))
        content.bind(minimum_height=content.setter('height'))

        stats_row = BoxLayout(size_hint_y=None, height=dp(120), spacing=dp(12))

        # 修复：应用新图标类型
        self.card_invested = self.create_stat_card('moneybag', 'Total\nInvested', '$0.00', (0.2, 0.6, 0.86, 1))
        self.card_market = self.create_stat_card('trend', 'Market\nValue', '$0.00', (0.15, 0.68, 0.38, 1))
        self.card_profit = self.create_stat_card('rocket', 'Total\nProfit', '$0.00', (0.61, 0.35, 0.71, 1))

        stats_row.add_widget(self.card_invested)
        stats_row.add_widget(self.card_market)
        stats_row.add_widget(self.card_profit)
        content.add_widget(stats_row)

        fire_card = ModernCard(orientation='vertical', size_hint_y=None, height=dp(180), spacing=dp(10))
        fire_title = Label(text='FIRE Progress', font_size='18sp', bold=True,
                           color=(0.2, 0.2, 0.2, 1), size_hint_y=None, height=dp(30))
        fire_card.add_widget(fire_title)

        ring_container = BoxLayout(size_hint_y=None, height=dp(120))
        self.fire_ring = ProgressRing(size_hint=(None, None), size=(dp(120), dp(120)))
        self.fire_ring.progress = 0

        ring_box = BoxLayout()
        ring_box.add_widget(BoxLayout())
        ring_box.add_widget(self.fire_ring)
        ring_box.add_widget(BoxLayout())
        ring_container.add_widget(ring_box)

        self.lbl_fire_pct = Label(text='0%', font_size='24sp', bold=True, color=(0.2, 0.6, 0.86, 1))
        ring_container.add_widget(self.lbl_fire_pct)
        fire_card.add_widget(ring_container)
        content.add_widget(fire_card)

        chart_card = ModernCard(orientation='vertical', size_hint_y=None, height=dp(280), spacing=dp(10))
        chart_title = Label(text='Asset Growth (Principal + Profit)', font_size='18sp', bold=True,
                            color=(0.2, 0.2, 0.2, 1), size_hint_y=None, height=dp(30))
        chart_card.add_widget(chart_title)

        self.investment_chart = SimpleLineChart(size_hint_y=None, height=dp(220))
        chart_card.add_widget(self.investment_chart)
        content.add_widget(chart_card)

        actions_card = ModernCard(orientation='vertical', size_hint_y=None, height=dp(140), spacing=dp(10))
        actions_title = Label(text='Quick Actions', font_size='18sp', bold=True,
                              color=(0.2, 0.2, 0.2, 1), size_hint_y=None, height=dp(30))
        actions_card.add_widget(actions_title)

        actions_grid = GridLayout(cols=2, spacing=dp(10), size_hint_y=None, height=dp(80))

        btn_invest = GradientButton(text='New Investment', font_size='14sp', bold=True,
                                    color1=[0.2, 0.6, 0.86, 1], color2=[0.15, 0.45, 0.72, 1])
        btn_invest.bind(on_press=lambda x: setattr(self.parent, 'current', 'invest'))

        btn_analyze = GradientButton(text='Analytics', font_size='14sp', bold=True,
                                     color1=[0.61, 0.35, 0.71, 1], color2=[0.51, 0.25, 0.61, 1])
        btn_analyze.bind(on_press=lambda x: setattr(self.parent, 'current', 'analytics'))

        btn_backup = GradientButton(text='Backup', font_size='14sp', bold=True,
                                    color1=[0.15, 0.68, 0.38, 1], color2=[0.12, 0.55, 0.30, 1])
        btn_backup.bind(on_press=self.do_backup)

        self.btn_refresh = GradientButton(text='Refresh', font_size='14sp', bold=True,
                                          color1=[0.95, 0.61, 0.07, 1], color2=[0.85, 0.51, 0.05, 1])
        self.btn_refresh.bind(on_press=self.refresh_dashboard)

        actions_grid.add_widget(btn_invest)
        actions_grid.add_widget(btn_analyze)
        actions_grid.add_widget(btn_backup)
        actions_grid.add_widget(self.btn_refresh)
        actions_card.add_widget(actions_grid)
        content.add_widget(actions_card)

        content.add_widget(BoxLayout(size_hint_y=None, height=dp(80)))
        scroll.add_widget(content)
        layout.add_widget(scroll)
        self.add_widget(layout)

        Clock.schedule_once(lambda dt: self.refresh_dashboard(None), 0.5)

    def create_stat_card(self, icon_type, title, value, bg_color):
        card = BoxLayout(orientation='vertical', spacing=dp(5), padding=dp(10))

        with card.canvas.before:
            Color(*bg_color)
            card.bg_rect = RoundedRectangle(pos=card.pos, size=card.size, radius=[dp(15)])

        card.bind(pos=lambda i, v: setattr(i.bg_rect, 'pos', v),
                  size=lambda i, v: setattr(i.bg_rect, 'size', v))

        # 使用 StatIcon 组件
        icon_widget = StatIcon(icon_type=icon_type, size_hint_y=0.4)
        card.add_widget(icon_widget)

        # 修复：设置 halign 和 valign 确保完全居中
        title_lbl = Label(
            text=title,
            font_size='11sp',
            size_hint_y=0.3,
            color=(1, 1, 1, 0.9),
            halign='center',  # 强制水平居中
            valign='middle'  # 强制垂直居中
        )
        # 必须绑定 text_size 到 size 才能使 halign/valign 生效
        title_lbl.bind(size=lambda s, w: setattr(s, 'text_size', s.size))
        card.add_widget(title_lbl)

        value_lbl = Label(
            text=value,
            font_size='18sp',
            bold=True,
            size_hint_y=0.3,
            color=(1, 1, 1, 1),
            halign='center',  # 同样对 value 应用居中，以防万一
            valign='middle'
        )
        value_lbl.bind(size=lambda s, w: setattr(s, 'text_size', s.size))
        card.add_widget(value_lbl)

        card.title_label = title_lbl
        card.value_label = value_lbl

        return card

    def refresh_dashboard(self, instance):
        if instance:
            instance.disabled = True
            instance.text = "Loading..."

        threading.Thread(target=self._refresh_thread, daemon=True).start()

    def _refresh_thread(self):
        ndx, prices, err = self.backend.get_market_data()

        self.backend.latest_prices = prices

        realtime_data = self.backend.get_avg_cost(current_prices=prices)
        chart_data = self.backend.get_chart_data()

        Clock.schedule_once(lambda dt: self._update_dashboard_ui(realtime_data, chart_data), 0)

    def _update_dashboard_ui(self, data, chart_data):
        self.btn_refresh.disabled = False
        self.btn_refresh.text = "Refresh"

        if data:
            invested = data.get('principal', 0)
            market_val = data.get('market_value', 0)
            profit = data.get('profit', 0)

            self.card_invested.value_label.text = f'${float(invested):,.2f}'
            self.card_market.value_label.text = f'${float(market_val):,.2f}'

            profit_text = f'${float(abs(profit)):,.2f}'
            if profit >= 0:
                self.card_profit.value_label.text = f'+{profit_text}'
            else:
                self.card_profit.value_label.text = f'-{profit_text}'

            goal = self.backend.config.get('fire_goal', 1000000)
            fire_pct = min((invested / goal) * 100, 100) if goal > 0 else 0

            self.fire_ring.progress = float(fire_pct)
            self.lbl_fire_pct.text = f'{float(fire_pct):.1f}%'
        else:
            self.card_invested.value_label.text = '$0.00'
            self.card_market.value_label.text = '$0.00'
            self.card_profit.value_label.text = '$0.00'
            self.fire_ring.progress = 0.0
            self.lbl_fire_pct.text = '0%'

        if chart_data and data:
            m_data = chart_data.get('market_value', [])
            if m_data:
                updated_data = m_data + [data['market_value']]
                self.investment_chart.set_data(updated_data)
            elif chart_data.get('investment'):
                i_data = chart_data.get('investment', [])
                updated_data = i_data + [data['principal']]
                self.investment_chart.set_data(updated_data)

    def do_backup(self, instance):
        if self.backend.backup_data():
            self.show_popup('Success', 'Data backed up successfully!')
        else:
            self.show_popup('Info', 'No data to backup')

    def show_popup(self, title, message):
        popup = Popup(title=title, content=Label(text=message, font_size='14sp'), size_hint=(0.8, 0.35))
        popup.open()


# ==================== Investment Screen ====================
class InvestmentScreen(Screen):
    def __init__(self, backend, **kwargs):
        super().__init__(**kwargs)
        self.backend = backend
        self.ndx = None
        self.prices = None
        self.strategy = None
        self.build_ui()

    def build_ui(self):
        layout = BoxLayout(orientation='vertical', spacing=0)

        scroll = ScrollView()
        content = BoxLayout(orientation='vertical', size_hint_y=None, spacing=dp(15), padding=dp(15))
        content.bind(minimum_height=content.setter('height'))

        status_card = ModernCard(orientation='horizontal', size_hint_y=None, height=dp(50), spacing=dp(10))
        self.lbl_bonus = Label(text='Bonus: $10,000.00', font_size='15sp', color=(0.2, 0.6, 0.86, 1), bold=True)
        status_card.add_widget(self.lbl_bonus)
        content.add_widget(status_card)

        market_card = ModernCard(orientation='vertical', size_hint_y=None, height=dp(400), spacing=dp(10))
        market_title = Label(text='Market Data', font_size='18sp', bold=True,
                             color=(0.2, 0.2, 0.2, 1), size_hint_y=None, height=dp(30))
        market_card.add_widget(market_title)

        self.btn_refresh = GradientButton(text='Refresh Market', size_hint_y=None, height=dp(50),
                                          font_size='15sp', bold=True, color1=[0.2, 0.6, 0.86, 1])
        self.btn_refresh.bind(on_press=self.refresh_data)
        market_card.add_widget(self.btn_refresh)

        info_grid = GridLayout(cols=2, size_hint_y=None, height=dp(100), spacing=dp(10))
        self.lbl_ndx = Label(text='NDX: --', font_size='13sp', color=(0.3, 0.3, 0.3, 1))
        self.lbl_ma120 = Label(text='MA120: --', font_size='13sp', color=(0.61, 0.35, 0.71, 1))
        self.lbl_rsi = Label(text='RSI: --', font_size='13sp', color=(0.91, 0.49, 0.13, 1))
        self.lbl_indicators = Label(text='Signals: --', font_size='13sp', color=(0.15, 0.68, 0.38, 1))

        info_grid.add_widget(self.lbl_ndx)
        info_grid.add_widget(self.lbl_ma120)
        info_grid.add_widget(self.lbl_rsi)
        info_grid.add_widget(self.lbl_indicators)
        market_card.add_widget(info_grid)

        valuation_box = BoxLayout(size_hint_y=None, height=dp(30), spacing=dp(10))
        valuation_box.add_widget(Label(text='Valuation:', font_size='13sp', color=(0.3, 0.3, 0.3, 1), size_hint_x=0.3))
        self.lbl_valuation = Label(text='--', font_size='13sp', bold=True, color=(0.5, 0.5, 0.5, 1), size_hint_x=0.7)
        valuation_box.add_widget(self.lbl_valuation)
        market_card.add_widget(valuation_box)

        self.valuation_bar = BoxLayout(size_hint_y=None, height=dp(25), padding=[dp(5), 0])
        market_card.add_widget(self.valuation_bar)

        self.lbl_advice = Label(text='Waiting...', font_size='15sp', bold=True,
                                color=(0.15, 0.68, 0.38, 1), size_hint_y=None, height=dp(50))
        market_card.add_widget(self.lbl_advice)
        content.add_widget(market_card)

        calc_card = ModernCard(orientation='vertical', size_hint_y=None, height=dp(160), spacing=dp(10))
        calc_title = Label(text='Investment Amount', font_size='18sp', bold=True,
                           color=(0.2, 0.2, 0.2, 1), size_hint_y=None, height=dp(30))
        calc_card.add_widget(calc_title)

        input_box = BoxLayout(size_hint_y=None, height=dp(50), spacing=dp(10))
        input_box.add_widget(Label(text='Amount ($):', size_hint_x=0.35, font_size='15sp',
                                   color=(0.3, 0.3, 0.3, 1), bold=True))
        self.entry_cash = StyledTextInput(text='', input_filter='float', size_hint_x=0.65, font_size='16sp')
        input_box.add_widget(self.entry_cash)
        calc_card.add_widget(input_box)

        self.btn_calc = GradientButton(text='Calculate Shares', size_hint_y=None, height=dp(50),
                                       disabled=True, font_size='15sp', bold=True,
                                       color1=[0.61, 0.35, 0.71, 1])
        self.btn_calc.bind(on_press=self.calc_plan)
        calc_card.add_widget(self.btn_calc)
        content.add_widget(calc_card)

        exec_card = ModernCard(orientation='vertical', size_hint_y=None, height=dp(250), spacing=dp(10))
        voo_label = Label(text='VOO (S&P 500)', font_size='14sp', bold=True,
                          color=(0.2, 0.6, 0.86, 1), size_hint_y=None, height=dp(25))
        exec_card.add_widget(voo_label)

        voo_box = BoxLayout(size_hint_y=None, height=dp(45), spacing=dp(8))
        voo_box.add_widget(Label(text='Price ($)', size_hint_x=0.25, font_size='13sp', color=(0.5, 0.5, 0.5, 1)))
        self.entry_price_a = StyledTextInput(text='', input_filter='float', size_hint_x=0.35)
        voo_box.add_widget(self.entry_price_a)
        voo_box.add_widget(Label(text='Shares', size_hint_x=0.15, font_size='13sp', color=(0.5, 0.5, 0.5, 1)))
        self.entry_share_a = StyledTextInput(text='', input_filter='float', size_hint_x=0.25)
        voo_box.add_widget(self.entry_share_a)
        exec_card.add_widget(voo_box)

        qqqm_label = Label(text='QQQM (NASDAQ)', font_size='14sp', bold=True,
                           color=(0.61, 0.35, 0.71, 1), size_hint_y=None, height=dp(25))
        exec_card.add_widget(qqqm_label)

        qqqm_box = BoxLayout(size_hint_y=None, height=dp(45), spacing=dp(8))
        qqqm_box.add_widget(Label(text='Price ($)', size_hint_x=0.25, font_size='13sp', color=(0.5, 0.5, 0.5, 1)))
        self.entry_price_b = StyledTextInput(text='', input_filter='float', size_hint_x=0.35)
        qqqm_box.add_widget(self.entry_price_b)
        qqqm_box.add_widget(Label(text='Shares', size_hint_x=0.15, font_size='13sp', color=(0.5, 0.5, 0.5, 1)))
        self.entry_share_b = StyledTextInput(text='', input_filter='float', size_hint_x=0.25)
        qqqm_box.add_widget(self.entry_share_b)
        exec_card.add_widget(qqqm_box)

        self.btn_save = GradientButton(text='Save Record', size_hint_y=None, height=dp(52),
                                       disabled=True, font_size='16sp', bold=True,
                                       color1=[0.18, 0.8, 0.44, 1])
        self.btn_save.bind(on_press=self.save_data)
        exec_card.add_widget(self.btn_save)
        content.add_widget(exec_card)

        action_card = ModernCard(orientation='vertical', size_hint_y=None, height=dp(140), spacing=dp(10))
        action_title = Label(text='Actions', font_size='18sp', bold=True,
                             color=(0.2, 0.2, 0.2, 1), size_hint_y=None, height=dp(30))
        action_card.add_widget(action_title)

        action_grid = GridLayout(cols=2, spacing=dp(10), size_hint_y=None, height=dp(80))

        btn_undo = GradientButton(text='Undo Last', font_size='14sp', bold=True,
                                  color1=[0.95, 0.61, 0.07, 1])
        btn_undo.bind(on_press=self.do_undo)

        btn_records = GradientButton(text='View Records', font_size='14sp', bold=True,
                                     color1=[0.2, 0.6, 0.86, 1])
        btn_records.bind(on_press=self.view_records)

        action_grid.add_widget(btn_undo)
        action_grid.add_widget(btn_records)
        action_card.add_widget(action_grid)
        content.add_widget(action_card)

        content.add_widget(BoxLayout(size_hint_y=None, height=dp(80)))
        scroll.add_widget(content)
        layout.add_widget(scroll)
        self.add_widget(layout)

    def on_enter(self):
        self.refresh_ui_status()

    def refresh_ui_status(self):
        self.backend.reload_status()
        remaining = self.backend.config.get('bonus_fund_remaining', 0)
        self.lbl_bonus.text = f'Bonus: ${float(remaining):,.2f}'

    def refresh_data(self, instance):
        self.btn_refresh.disabled = True
        self.btn_refresh.text = 'Loading...'
        threading.Thread(target=self._refresh_thread, daemon=True).start()

    def _refresh_thread(self):
        self.ndx, self.prices, err = self.backend.get_market_data()
        Clock.schedule_once(lambda dt: self._update_ui(err), 0)

    def _update_ui(self, err):
        self.btn_refresh.disabled = False
        self.btn_refresh.text = 'Refresh Market'

        if self.ndx:
            self.strategy = self.backend.calculate_strategy(self.ndx)

            change_color = (0.15, 0.68, 0.38, 1) if self.ndx['change'] >= 0 else (0.91, 0.3, 0.24, 1)
            self.lbl_ndx.text = f"NDX: {self.ndx['close']:,.2f} ({self.ndx['change']:+.2f}%)"
            self.lbl_ndx.color = change_color

            self.lbl_ma120.text = f"MA120: {self.ndx['ma120']:,.2f}"
            self.lbl_rsi.text = f"RSI: {self.ndx['rsi']:.1f}"

            ind_count = self.strategy['indicator_count']
            if ind_count > 0:
                self.lbl_indicators.text = f"{ind_count} Signals: " + ",".join(self.strategy['indicator_details'])
            else:
                self.lbl_indicators.text = "No signals"

            percentile = self.ndx.get('valuation_percentile', 50.0)
            history_count = self.ndx.get('valuation_history_count', 0)

            level, color = self.backend.get_valuation_level(percentile)

            if history_count > 0:
                self.lbl_valuation.text = f"{percentile:.1f}% ({level}) [Since 2015 {history_count}d]"
            else:
                self.lbl_valuation.text = f"{percentile:.1f}% ({level}) [No Data]"
            self.lbl_valuation.color = color

            self.valuation_bar.canvas.before.clear()
            with self.valuation_bar.canvas.before:
                total_blocks = 30
                block_spacing = dp(2)

                total_spacing = (total_blocks - 1) * block_spacing
                available_width = self.valuation_bar.width - total_spacing
                block_width = available_width / total_blocks
                block_height = self.valuation_bar.height - dp(4)

                filled_blocks = int((percentile / 100) * total_blocks)

                for i in range(total_blocks):
                    x = self.valuation_bar.x + i * (block_width + block_spacing)
                    y = self.valuation_bar.y + dp(2)

                    if i < filled_blocks:
                        Color(*color)
                    else:
                        Color(0.85, 0.85, 0.85, 1)

                    RoundedRectangle(pos=(x, y), size=(block_width, block_height), radius=[dp(2)])

            reg = self.strategy['regular_amount']
            bonus = self.strategy['bonus_amount']
            total = self.strategy['total_amount']

            if bonus > 0:
                self.lbl_advice.text = f"Suggested: ${total:.0f} (${reg}+${bonus:.0f})"
            else:
                self.lbl_advice.text = f"Suggested: ${reg} (regular)"

            self.entry_cash.text = str(int(total))

            sa = self.backend.config['symbol_a']
            sb = self.backend.config['symbol_b']

            if sa in self.prices:
                self.entry_price_a.text = f"{self.prices[sa]:.2f}"
            if sb in self.prices:
                self.entry_price_b.text = f"{self.prices[sb]:.2f}"

            self.btn_calc.disabled = False
            self.refresh_ui_status()

    def calc_plan(self, instance):
        try:
            total = float(self.entry_cash.text)
            pa = float(self.entry_price_a.text)
            pb = float(self.entry_price_b.text)

            target = total / 2
            sa = round(target / pa, 4) if pa > 0 else 0
            sb = round(target / pb, 4) if pb > 0 else 0

            self.entry_share_a.text = str(sa)
            self.entry_share_b.text = str(sb)

            self.btn_save.disabled = False
            self.show_popup('Success', f'VOO: {sa}\nQQQM: {sb}')
        except:
            self.show_popup('Error', 'Invalid input')

    def save_data(self, instance):
        try:
            details = {
                'price_a': float(self.entry_price_a.text),
                'share_a': float(self.entry_share_a.text),
                'price_b': float(self.entry_price_b.text),
                'share_b': float(self.entry_share_b.text),
            }

            self.backend.save_transaction(self.ndx, self.strategy, details)
            self.refresh_ui_status()

            msg = f"Saved!\n\nRegular: ${self.strategy['regular_amount']}\nBonus: ${self.strategy['bonus_amount']:.2f}"
            self.show_popup('Success', msg)

            self.btn_save.disabled = True

            if hasattr(self.parent, 'get_screen'):
                dash = self.parent.get_screen('dashboard')
                if hasattr(dash, 'refresh_dashboard'):
                    dash.refresh_dashboard(None)
        except Exception as e:
            self.show_popup('Error', f'Failed: {str(e)}')

    def do_undo(self, instance):
        suc, msg = self.backend.undo_last_transaction()
        if suc:
            self.refresh_ui_status()
            if hasattr(self.parent, 'get_screen'):
                dash = self.parent.get_screen('dashboard')
                if hasattr(dash, 'refresh_dashboard'):
                    dash.refresh_dashboard(None)
        self.show_popup('Undo', msg)

    def view_records(self, instance):
        df = self.backend.get_all_records()
        if df is None or len(df) == 0:
            self.show_popup('No Records', 'No data yet!')
            return

        content = BoxLayout(orientation='vertical', padding=dp(15), spacing=dp(10))

        header = BoxLayout(size_hint_y=None, height=dp(40), spacing=dp(10))
        header.add_widget(Label(
            text=f'📊 {len(df)} Investment Records',
            font_size='18sp',
            bold=True,
            color=(0, 0.45, 0.73, 1)
        ))
        content.add_widget(header)

        scroll = ScrollView()
        records_layout = BoxLayout(orientation='vertical', size_hint_y=None, spacing=dp(12))
        records_layout.bind(minimum_height=records_layout.setter('height'))

        try:
            import yfinance as yf
            current_voo_price = yf.Ticker(self.backend.voo_ticker).history(period='1d')['Close'].iloc[-1]
            current_qqqm_price = yf.Ticker(self.backend.qqqm_ticker).history(period='1d')['Close'].iloc[-1]
            use_realtime = True
        except:
            use_realtime = False

        for idx, row in df.tail(10).iterrows():
            record_card = BoxLayout(orientation='vertical', size_hint_y=None, height=dp(165), padding=dp(15),
                                    spacing=dp(5))

            with record_card.canvas.before:
                Color(1, 1, 1, 1)
                record_card.rect = RoundedRectangle(pos=record_card.pos, size=record_card.size, radius=[dp(8)])
                Color(0.9, 0.9, 0.9, 0.3)
                record_card.shadow = RoundedRectangle(
                    pos=(record_card.x + dp(1), record_card.y - dp(1)),
                    size=record_card.size,
                    radius=[dp(8)]
                )

            def update_card_bg(instance, value):
                instance.rect.pos = instance.pos
                instance.rect.size = instance.size
                instance.shadow.pos = (instance.x + dp(1), instance.y - dp(1))
                instance.shadow.size = instance.size

            record_card.bind(pos=update_card_bg, size=update_card_bg)

            date_str = str(row.get('Date', ''))[:10]
            ndx_change = str(row.get('NDX_Change', '0%'))

            row1 = BoxLayout(size_hint_y=None, height=dp(25))

            # Removed Emoji 📅
            date_label = Label(
                text=f"Date: {date_str}",
                font_size='14sp',
                bold=True,
                color=(0.1, 0.1, 0.1, 1)
            )
            date_label.text_size = (date_label.width, None)
            date_label.halign = 'left'
            date_label.bind(size=lambda i, v: setattr(i, 'text_size', (i.width, None)))
            row1.add_widget(date_label)

            change_color = (0.13, 0.7, 0.38, 1) if '-' not in ndx_change else (0.9, 0.3, 0.24, 1)
            ndx_label = Label(
                text=f"NDX {ndx_change}",
                font_size='12sp',
                color=change_color,
                bold=True
            )
            ndx_label.text_size = (ndx_label.width, None)
            ndx_label.halign = 'right'
            ndx_label.bind(size=lambda i, v: setattr(i, 'text_size', (i.width, None)))
            row1.add_widget(ndx_label)

            record_card.add_widget(row1)

            regular = float(row.get('Regular', 0))
            bonus = float(row.get('Bonus', 0))
            total_cost = float(row.get('Monthly_Cost', 0))

            row2 = BoxLayout(size_hint_y=None, height=dp(24))
            # Removed Emoji 💰
            if bonus > 0:
                amount_text = f"Cost: ${total_cost:,.2f} = Reg ${regular:,.0f} + Bonus ${bonus:,.0f}"
            else:
                amount_text = f"Cost: Reg ${regular:,.0f}"

            amount_label = Label(
                text=amount_text,
                font_size='12sp',
                color=(0.4, 0.4, 0.4, 1)
            )
            amount_label.text_size = (amount_label.width, None)
            amount_label.halign = 'left'
            amount_label.bind(size=lambda i, v: setattr(i, 'text_size', (i.width, None)))
            row2.add_widget(amount_label)
            record_card.add_widget(row2)

            voo_price = float(row.get('VOO_Price', 0))
            voo_shares = float(row.get('VOO_Shares', 0))

            row3 = BoxLayout(size_hint_y=None, height=dp(24))

            # Removed Emoji 🔵
            voo_name = Label(
                text=f"VOO",
                font_size='12sp',
                color=(0, 0.45, 0.73, 1),
                size_hint_x=0.3
            )
            voo_name.text_size = (voo_name.width, None)
            voo_name.halign = 'left'
            voo_name.bind(size=lambda i, v: setattr(i, 'text_size', (i.width, None)))
            row3.add_widget(voo_name)

            voo_detail = Label(
                text=f"${voo_price:.2f} x {voo_shares:.3f} sh",
                font_size='11sp',
                color=(0.3, 0.3, 0.3, 1)
            )
            voo_detail.text_size = (voo_detail.width, None)
            voo_detail.halign = 'right'
            voo_detail.bind(size=lambda i, v: setattr(i, 'text_size', (i.width, None)))
            row3.add_widget(voo_detail)

            record_card.add_widget(row3)

            qqqm_price = float(row.get('QQQM_Price', 0))
            qqqm_shares = float(row.get('QQQM_Shares', 0))

            row4 = BoxLayout(size_hint_y=None, height=dp(24))

            # Removed Emoji 🟣
            qqqm_name = Label(
                text=f"QQQM",
                font_size='12sp',
                color=(0.4, 0.28, 0.6, 1),
                size_hint_x=0.3
            )
            qqqm_name.text_size = (qqqm_name.width, None)
            qqqm_name.halign = 'left'
            qqqm_name.bind(size=lambda i, v: setattr(i, 'text_size', (i.width, None)))
            row4.add_widget(qqqm_name)

            qqqm_detail = Label(
                text=f"${qqqm_price:.2f} x {qqqm_shares:.3f} sh",
                font_size='11sp',
                color=(0.3, 0.3, 0.3, 1)
            )
            qqqm_detail.text_size = (qqqm_detail.width, None)
            qqqm_detail.halign = 'right'
            qqqm_detail.bind(size=lambda i, v: setattr(i, 'text_size', (i.width, None)))
            row4.add_widget(qqqm_detail)

            record_card.add_widget(row4)

            if use_realtime:
                current_market_value = (voo_shares * current_voo_price +
                                        qqqm_shares * current_qqqm_price)
                current_profit = current_market_value - total_cost
                market_val = current_market_value
                profit = current_profit
            else:
                market_val = float(row.get('Market_Value', 0))
                profit = float(row.get('Profit', 0))

            row5 = BoxLayout(size_hint_y=None, height=dp(26))

            # Removed Emoji 💵
            market_label = Label(
                text=f"Val: ${market_val:,.2f}",
                font_size='13sp',
                color=(0.3, 0.3, 0.3, 1),
                bold=True
            )
            market_label.text_size = (market_label.width, None)
            market_label.halign = 'left'
            market_label.bind(size=lambda i, v: setattr(i, 'text_size', (i.width, None)))
            row5.add_widget(market_label)

            profit_color = (0.13, 0.7, 0.38, 1) if profit >= 0 else (0.9, 0.3, 0.24, 1)
            profit_sign = '+' if profit >= 0 else ''
            profit_label = Label(
                text=f"{profit_sign}${profit:,.2f}",
                font_size='14sp',
                bold=True,
                color=profit_color
            )
            profit_label.text_size = (profit_label.width, None)
            profit_label.halign = 'right'
            profit_label.bind(size=lambda i, v: setattr(i, 'text_size', (i.width, None)))
            row5.add_widget(profit_label)

            record_card.add_widget(row5)

            records_layout.add_widget(record_card)

        scroll.add_widget(records_layout)
        content.add_widget(scroll)

        btn_close = GradientButton(text='Close', size_hint_y=None, height=dp(48), color1=[0.6, 0.6, 0.6, 1])
        popup = Popup(title='Investment Records', content=content, size_hint=(0.9, 0.85))
        btn_close.bind(on_press=popup.dismiss)
        content.add_widget(btn_close)

        popup.open()

    def show_popup(self, title, message):
        popup = Popup(title=title, content=Label(text=message, font_size='14sp'), size_hint=(0.85, 0.4))
        popup.open()


# ==================== Analytics Screen ====================
class AnalyticsScreen(Screen):
    def __init__(self, backend, **kwargs):
        super().__init__(**kwargs)
        self.backend = backend
        self.build_ui()

    def build_ui(self):
        layout = BoxLayout(orientation='vertical', spacing=0)

        scroll = ScrollView()
        content = BoxLayout(orientation='vertical', size_hint_y=None, spacing=dp(15), padding=dp(15))
        content.bind(minimum_height=content.setter('height'))

        summary_card = ModernCard(orientation='vertical', size_hint_y=None, height=dp(200), spacing=dp(10))
        summary_card.add_widget(Label(text='Portfolio Summary', font_size='18sp', bold=True,
                                      color=(0.2, 0.2, 0.2, 1), size_hint_y=None, height=dp(30)))

        self.summary_grid = GridLayout(cols=2, spacing=dp(10))
        summary_card.add_widget(self.summary_grid)
        content.add_widget(summary_card)

        holdings_card = ModernCard(orientation='vertical', size_hint_y=None, height=dp(220), spacing=dp(10))
        holdings_card.add_widget(Label(text='Holdings', font_size='18sp', bold=True,
                                       color=(0.2, 0.2, 0.2, 1), size_hint_y=None, height=dp(30)))

        self.holdings_layout = BoxLayout(orientation='vertical', spacing=dp(10))
        holdings_card.add_widget(self.holdings_layout)
        content.add_widget(holdings_card)

        perf_card = ModernCard(orientation='vertical', size_hint_y=None, height=dp(280), spacing=dp(10))

        # 标题修改：强调是 Cost vs Market Value
        perf_card.add_widget(Label(text='Cost vs Market Value', font_size='18sp', bold=True,
                                   color=(0.2, 0.2, 0.2, 1), size_hint_y=None, height=dp(30)))

        # 这里替换为 DoubleLineChart
        self.perf_chart = DoubleLineChart(size_hint_y=None, height=dp(220))
        perf_card.add_widget(self.perf_chart)
        content.add_widget(perf_card)

        content.add_widget(BoxLayout(size_hint_y=None, height=dp(80)))
        scroll.add_widget(content)
        layout.add_widget(scroll)
        self.add_widget(layout)

    def on_enter(self):
        self.refresh_analytics()

    def refresh_analytics(self):
        # 核心修复1：尝试使用 backend.latest_prices，确保与 Dashboard 刷新后的数据一致
        data = self.backend.get_avg_cost(current_prices=self.backend.latest_prices)

        self.summary_grid.clear_widgets()

        if data:
            invested = float(data.get('principal', 0))
            market_val = float(data.get('market_value', 0))
            profit = float(data.get('profit', 0))

            self.summary_grid.add_widget(Label(text='Total Invested:', color=(0.3, 0.3, 0.3, 1)))
            # 修复2：显示两位小数
            self.summary_grid.add_widget(Label(text=f'${invested:,.2f}', bold=True, color=(0.2, 0.6, 0.86, 1)))
            self.summary_grid.add_widget(Label(text='Market Value:', color=(0.3, 0.3, 0.3, 1)))
            self.summary_grid.add_widget(Label(text=f'${market_val:,.2f}', bold=True, color=(0.15, 0.68, 0.38, 1)))
            self.summary_grid.add_widget(Label(text='Total Profit:', color=(0.3, 0.3, 0.3, 1)))
            profit_color = (0.15, 0.68, 0.38, 1) if profit >= 0 else (0.91, 0.3, 0.24, 1)
            self.summary_grid.add_widget(Label(text=f'${profit:,.2f}', bold=True, color=profit_color))

            self.holdings_layout.clear_widgets()

            sym_a = self.backend.config['symbol_a']
            sym_b = self.backend.config['symbol_b']

            if sym_a in data:
                avg_a, shares_a = data[sym_a]
                voo_box = BoxLayout(orientation='vertical', padding=dp(10))
                with voo_box.canvas.before:
                    Color(0.2, 0.6, 0.86, 0.2)
                    voo_box.rect = RoundedRectangle(pos=voo_box.pos, size=voo_box.size, radius=[dp(10)])
                voo_box.bind(pos=lambda i, v: setattr(i.rect, 'pos', v),
                             size=lambda i, v: setattr(i.rect, 'size', v))

                voo_box.add_widget(Label(text=f'{sym_a}', font_size='16sp', bold=True, color=(0.2, 0.6, 0.86, 1)))
                voo_box.add_widget(Label(text=f'{shares_a:.2f} shares @ ${avg_a:.2f}',
                                         font_size='14sp', color=(0.3, 0.3, 0.3, 1)))
                self.holdings_layout.add_widget(voo_box)

            if sym_b in data:
                avg_b, shares_b = data[sym_b]
                qqqm_box = BoxLayout(orientation='vertical', padding=dp(10))
                with qqqm_box.canvas.before:
                    Color(0.61, 0.35, 0.71, 0.2)
                    qqqm_box.rect = RoundedRectangle(pos=qqqm_box.pos, size=qqqm_box.size, radius=[dp(10)])
                qqqm_box.bind(pos=lambda i, v: setattr(i.rect, 'pos', v),
                              size=lambda i, v: setattr(i.rect, 'size', v))

                qqqm_box.add_widget(Label(text=f'{sym_b}', font_size='16sp', bold=True, color=(0.61, 0.35, 0.71, 1)))
                qqqm_box.add_widget(Label(text=f'{shares_b:.2f} shares @ ${avg_b:.2f}',
                                          font_size='14sp', color=(0.3, 0.3, 0.3, 1)))
                self.holdings_layout.add_widget(qqqm_box)

        # 更新图表数据：传入 Market Value 和 Investment 两个列表
        chart_data = self.backend.get_chart_data()
        if chart_data:
            m_val = chart_data.get('market_value', [])
            i_val = chart_data.get('investment', [])
            self.perf_chart.set_data(market_data=m_val, cost_data=i_val)


# ==================== Settings Screen (V23.17 修复版) ====================
class SettingsScreen(Screen):
    def __init__(self, backend, **kwargs):
        super().__init__(**kwargs)
        self.backend = backend
        self.build_ui()

    def build_ui(self):
        layout = BoxLayout(orientation='vertical', spacing=0)

        # effect_cls='ScrollEffect' 去除过度回弹的物理效果，让滚动更跟手
        self.scroll = ScrollView(size_hint=(1, 1), effect_cls='ScrollEffect')
        
        content = BoxLayout(orientation='vertical', size_hint_y=None, spacing=dp(20), padding=[dp(15), dp(20), dp(15), dp(80)])
        content.bind(minimum_height=content.setter('height'))

        # --- 1. Configuration Card ---
        config_card = ModernCard(orientation='vertical', size_hint_y=None, height=dp(550), spacing=dp(5))
        config_card.add_widget(Label(text='Configuration', font_size='19sp', bold=True,
                                    color=(0.2,0.2,0.2,1), size_hint_y=None, height=dp(40)))

        # Symbol A
        config_card.add_widget(BoxLayout(size_hint_y=None, height=dp(5)))
        config_card.add_widget(Label(text='Symbol A', size_hint_y=None, height=dp(25), 
                                    color=(0.3,0.3,0.3,1), font_size='15sp', bold=True))
        self.entry_a = StyledTextInput(text=self.backend.config['symbol_a'])
        config_card.add_widget(self.entry_a)

        # Symbol B
        config_card.add_widget(BoxLayout(size_hint_y=None, height=dp(10)))
        config_card.add_widget(Label(text='Symbol B', size_hint_y=None, height=dp(25),
                                    color=(0.3,0.3,0.3,1), font_size='15sp', bold=True))
        self.entry_b = StyledTextInput(text=self.backend.config['symbol_b'])
        config_card.add_widget(self.entry_b)

        # FIRE Goal
        config_card.add_widget(BoxLayout(size_hint_y=None, height=dp(10)))
        config_card.add_widget(Label(text='FIRE Goal ($)', size_hint_y=None, height=dp(25),
                                    color=(0.3,0.3,0.3,1), font_size='15sp', bold=True))
        self.entry_fire = StyledTextInput(
            text=str(int(self.backend.config.get('fire_goal', 1000000))),
            input_filter='float'
        )
        config_card.add_widget(self.entry_fire)

        # Bonus Budget
        config_card.add_widget(BoxLayout(size_hint_y=None, height=dp(10)))
        config_card.add_widget(Label(text='Bonus Budget ($)', size_hint_y=None, height=dp(25),
                                    color=(0.3,0.3,0.3,1), font_size='15sp', bold=True))
        self.entry_bonus = StyledTextInput(
            text=str(int(self.backend.config.get('bonus_fund_total', 10000))),
            input_filter='float'
        )
        config_card.add_widget(self.entry_bonus)

        # Auto Refresh
        config_card.add_widget(BoxLayout(size_hint_y=None, height=dp(10)))
        config_card.add_widget(Label(text='Auto Refresh (min)', size_hint_y=None, height=dp(25),
                                    color=(0.3,0.3,0.3,1), font_size='15sp', bold=True))
        self.entry_refresh = StyledTextInput(
            text=str(int(self.backend.config.get('auto_refresh_interval', 60))),
            input_filter='int'
        )
        config_card.add_widget(self.entry_refresh)

        config_card.add_widget(BoxLayout(size_hint_y=None, height=dp(20)))
        btn_save = GradientButton(text='Save Settings', size_hint_y=None, height=dp(50),
                                 font_size='15sp', bold=True, color1=[0.18,0.8,0.44,1])
        btn_save.bind(on_press=self.save_settings)
        config_card.add_widget(btn_save)

        content.add_widget(config_card)

        # --- 2. Security Card ---
        # 修复：移除了之前的 Widget 占位符，现在标题位置会恢复正常（靠上，与 Configuration 对齐）
        security_card = ModernCard(orientation='vertical', size_hint_y=None, height=dp(240), spacing=dp(5))
        
        security_card.add_widget(Label(text='Security', font_size='19sp', bold=True,
                                      color=(0.2,0.2,0.2,1), size_hint_y=None, height=dp(40)))

        security_card.add_widget(BoxLayout(size_hint_y=None, height=dp(5)))
        security_card.add_widget(Label(text='Reset Password', size_hint_y=None, height=dp(25),
                                      color=(0.91,0.3,0.24,1), font_size='14sp', bold=True))
        self.entry_pwd = StyledTextInput(text='', password=True)
        security_card.add_widget(self.entry_pwd)

        security_card.add_widget(Label(text='(Leave blank to keep current)', size_hint_y=None, height=dp(22),
                                      color=(0.6,0.6,0.6,1), font_size='11sp'))

        btn_save_pwd = GradientButton(text='Update Password', size_hint_y=None, height=dp(48),
                                     font_size='15sp', bold=True, color1=[0.91,0.3,0.24,1])
        btn_save_pwd.bind(on_press=self.save_password)
        security_card.add_widget(btn_save_pwd)

        content.add_widget(security_card)

        # --- 3. Danger Zone Card ---
        danger_card = ModernCard(orientation='vertical', size_hint_y=None, height=dp(160), spacing=dp(5))
        danger_card.add_widget(Label(text='Danger Zone', font_size='19sp', bold=True,
                                    color=(0.91,0.3,0.24,1), size_hint_y=None, height=dp(40)))

        btn_reset = GradientButton(text='Reset All Data', size_hint_y=None, height=dp(50),
                                  font_size='15sp', bold=True, color1=[0.91,0.3,0.24,1])
        btn_reset.bind(on_press=self.do_reset)
        danger_card.add_widget(btn_reset)

        danger_card.add_widget(Label(text='This will delete all records!', 
                                    size_hint_y=None, height=dp(30),
                                    color=(0.6,0.6,0.6,1), font_size='12sp'))

        content.add_widget(danger_card)

        self.scroll.add_widget(content)
        layout.add_widget(self.scroll)
        self.add_widget(layout)

    def on_enter(self):
        # 修复2：使用 0.2 秒延时，确保手机渲染完布局后再执行置顶操作
        # 这能有效解决“进入页面位置不对”或“回弹”的问题
        Clock.schedule_once(self._force_scroll_top, 0.2)

    def _force_scroll_top(self, dt):
        self.scroll.scroll_y = 1
        self.scroll.update_from_scroll()

    def save_settings(self, instance):
        self.backend.config['symbol_a'] = self.entry_a.text.upper()
        self.backend.config['symbol_b'] = self.entry_b.text.upper()
        try:
            self.backend.config['fire_goal'] = float(self.entry_fire.text)
            self.backend.config['bonus_fund_total'] = float(self.entry_bonus.text)
            interval = int(self.entry_refresh.text)
            if interval < 1: interval = 1
            self.backend.config['auto_refresh_interval'] = interval
            
            app = App.get_running_app()
            if app:
                app.start_auto_refresh_timer()
                
        except:
            pass

        self.backend.save_config()
        self.show_popup('Success', 'Settings saved!')

    def save_password(self, instance):
        if self.entry_pwd.text:
            self.backend.set_reset_password(self.entry_pwd.text)
            self.entry_pwd.text = ''
            self.show_popup('Success', 'Password updated!')
        else:
            self.show_popup('Info', 'No password entered')

    def do_reset(self, instance):
        content = BoxLayout(orientation='vertical', padding=dp(15), spacing=dp(12))

        content.add_widget(Label(text='WARNING\n\nDelete ALL records?\nThis CANNOT be undone!',
                                size_hint_y=0.4, color=(0.91,0.3,0.24,1), font_size='16sp', bold=True))

        content.add_widget(Label(text='Enter password:', size_hint_y=0.15, color=(0.3,0.3,0.3,1)))
        entry_password = StyledTextInput(text='', password=True, size_hint_y=0.2)
        content.add_widget(entry_password)

        content.add_widget(Label(text='Default: 1234\n(Change in Settings)', size_hint_y=0.15,
                                color=(0.6,0.6,0.6,1), font_size='11sp'))

        popup = Popup(title='Confirm Reset', content=content, size_hint=(0.9, 0.55))

        btn_box = BoxLayout(size_hint_y=0.2, spacing=dp(10))

        btn_confirm = GradientButton(text='Confirm', color1=[0.91,0.3,0.24,1])
        btn_cancel = GradientButton(text='Cancel', color1=[0.6,0.6,0.6,1])

        def confirm_reset(inst):
            if self.backend.verify_reset_password(entry_password.text):
                if self.backend.reset_data():
                    popup.dismiss()
                    self.show_popup('Success', 'Data reset!')

                    if hasattr(self.parent, 'get_screen'):
                        dash = self.parent.get_screen('dashboard')
                        if hasattr(dash, 'refresh_dashboard'):
                            dash.refresh_dashboard(None)
            else:
                self.show_popup('Wrong Password', 'Incorrect!')
                entry_password.text = ''

        btn_confirm.bind(on_press=confirm_reset)
        btn_cancel.bind(on_press=popup.dismiss)

        btn_box.add_widget(btn_confirm)
        btn_box.add_widget(btn_cancel)
        content.add_widget(btn_box)

        popup.open()

    def show_popup(self, title, message):
        popup = Popup(title=title, content=Label(text=message, font_size='14sp'), size_hint=(0.8, 0.35))
        popup.open()
        

# ==================== Bottom Navigation Bar (Canvas Drawn Icons) ====================
class NavIcon(Widget):
    """
    Custom widget that draws icons using Canvas instructions to avoid font/unicode issues.
    """

    def __init__(self, icon_type, color=(0.5, 0.5, 0.5, 1), **kwargs):
        super().__init__(**kwargs)
        self.icon_type = icon_type
        self.color = color
        self.bind(pos=self.update_canvas, size=self.update_canvas)

    def set_color(self, color):
        self.color = color
        self.update_canvas()

    def update_canvas(self, *args):
        self.canvas.clear()
        cx, cy = self.center_x, self.center_y
        w, h = self.width, self.height

        # Define a drawing area that fits within the widget
        size = min(w, h) * 0.55

        with self.canvas:
            Color(*self.color)

            if self.icon_type == 'bars':  # Dashboard (Accounts style)
                # Draw 3 vertical bars of different heights
                bar_w = size / 3.5
                spacing = size / 10
                start_x = cx - (3 * bar_w + 2 * spacing) / 2
                bottom_y = cy - size / 2

                # Bar 1 (Medium)
                Rectangle(pos=(start_x, bottom_y), size=(bar_w, size * 0.6))
                # Bar 2 (Tall)
                Rectangle(pos=(start_x + bar_w + spacing, bottom_y), size=(bar_w, size))
                # Bar 3 (Short)
                Rectangle(pos=(start_x + 2 * (bar_w + spacing), bottom_y), size=(bar_w, size * 0.4))

            elif self.icon_type == 'arrows':  # Invest (Trade style)
                # Draw Up and Down arrows
                arrow_h = size
                arrow_w = size / 2.5
                offset_x = size / 4

                # Up Arrow (Left)
                up_x = cx - offset_x
                up_y_start = cy - arrow_h / 2
                up_y_end = cy + arrow_h / 2

                Line(points=[up_x, up_y_start, up_x, up_y_end], width=dp(1.5))
                Line(points=[up_x - arrow_w / 2, up_y_end - arrow_w / 2, up_x, up_y_end, up_x + arrow_w / 2,
                             up_y_end - arrow_w / 2], width=dp(1.5))

                # Down Arrow (Right)
                down_x = cx + offset_x
                down_y_start = cy + arrow_h / 2
                down_y_end = cy - arrow_h / 2

                Line(points=[down_x, down_y_start, down_x, down_y_end], width=dp(1.5))
                Line(points=[down_x - arrow_w / 2, down_y_end + arrow_w / 2, down_x, down_y_end, down_x + arrow_w / 2,
                             down_y_end + arrow_w / 2], width=dp(1.5))

            elif self.icon_type == 'globe':  # Analytics (Markets style)
                # Draw a simplified globe (Circle + Equator + Meridian)
                radius = size / 2
                Line(circle=(cx, cy, radius), width=dp(1.2))
                # Equator
                Line(points=[cx - radius, cy, cx + radius, cy], width=dp(1))
                # Meridian (Vertical Ellipse approximation)
                Line(ellipse=(cx - radius / 2.5, cy - radius, radius / 1.25, radius * 2), width=dp(1))

            elif self.icon_type == 'dots':  # Settings (More style)
                # Draw 3 horizontal dots
                dot_size = size / 4
                spacing = size / 4

                # Center Dot
                Ellipse(pos=(cx - dot_size / 2, cy - dot_size / 2), size=(dot_size, dot_size))
                # Left Dot
                Ellipse(pos=(cx - dot_size / 2 - dot_size - spacing, cy - dot_size / 2), size=(dot_size, dot_size))
                # Right Dot
                Ellipse(pos=(cx - dot_size / 2 + dot_size + spacing, cy - dot_size / 2), size=(dot_size, dot_size))


class NavTab(ButtonBehavior, BoxLayout):
    def __init__(self, text, icon_type, tab_id, screen_manager, **kwargs):
        super().__init__(**kwargs)
        self.orientation = 'vertical'
        self.tab_id = tab_id
        self.screen_manager = screen_manager

        self.padding = [0, dp(6), 0, dp(4)]
        self.spacing = dp(0)

        self.icon_widget = NavIcon(icon_type=icon_type, size_hint_y=0.65)
        self.text_label = Label(text=text, font_size='11sp', size_hint_y=0.35)

        self.add_widget(self.icon_widget)
        self.add_widget(self.text_label)

    def on_press(self):
        self.screen_manager.current = self.tab_id
        self.parent.update_active(self.tab_id)

    def set_active(self, is_active, color):
        if is_active:
            self.icon_widget.set_color(color)
            self.text_label.color = color
            self.text_label.bold = True
        else:
            gray = (0.6, 0.6, 0.6, 1)
            self.icon_widget.set_color(gray)
            self.text_label.color = gray
            self.text_label.bold = False


class BottomNavBar(BoxLayout):
    def __init__(self, screen_manager, **kwargs):
        super().__init__(**kwargs)
        self.orientation = 'horizontal'
        self.size_hint_y = None
        self.height = dp(65)
        self.screen_manager = screen_manager

        with self.canvas.before:
            Color(1, 1, 1, 1)
            self.bg_rect = Rectangle(pos=self.pos, size=self.size)
        self.bind(pos=lambda i, v: setattr(self.bg_rect, 'pos', v),
                  size=lambda i, v: setattr(self.bg_rect, 'size', v))

        # Add top border line
        with self.canvas.after:
            Color(0.9, 0.9, 0.9, 1)
            self.top_line = Line(points=[self.x, self.y + self.height, self.x + self.width, self.y + self.height],
                                 width=1)

        def update_line(instance, value):
            instance.top_line.points = [instance.x, instance.y + instance.height, instance.x + instance.width,
                                        instance.y + instance.height]

        self.bind(pos=update_line, size=update_line)

        self.nav_tabs = {}

        tabs = [
            ('dashboard', 'Dashboard', 'bars', [0.15, 0.55, 0.82, 1]),
            ('invest', 'Invest', 'arrows', [0.61, 0.35, 0.71, 1]),
            ('analytics', 'Analytics', 'globe', [0.15, 0.68, 0.38, 1]),
            ('settings', 'Settings', 'dots', [0.95, 0.61, 0.07, 1])
        ]

        for tab_id, text, icon_type, color in tabs:
            tab = NavTab(text, icon_type, tab_id, screen_manager)
            tab.tab_color = color
            self.nav_tabs[tab_id] = tab
            self.add_widget(tab)

        self.update_active('dashboard')

    def update_active(self, active_id):
        for tab_id, tab in self.nav_tabs.items():
            if tab_id == active_id:
                tab.set_active(True, tab.tab_color)
            else:
                tab.set_active(False, None)


# ==================== Main App ====================
class InvestmentProApp(App):
    def build(self):
        Window.clearcolor = (0.95, 0.96, 0.97, 1)

        self.backend = InvestmentBackend()
        self.refresh_event = None

        self.root_layout = BoxLayout(orientation='vertical')

        self.sm = ScreenManager()
        self.sm.add_widget(DashboardScreen(self.backend, name='dashboard'))
        self.sm.add_widget(InvestmentScreen(self.backend, name='invest'))
        self.sm.add_widget(AnalyticsScreen(self.backend, name='analytics'))
        self.sm.add_widget(SettingsScreen(self.backend, name='settings'))

        self.root_layout.add_widget(self.sm)

        nav = BottomNavBar(self.sm)
        self.root_layout.add_widget(nav)

        self.start_auto_refresh_timer()

        return self.root_layout

    def on_start(self):
        # 1. 显示开屏欢迎语
        self.show_welcome_message()

    def show_welcome_message(self):
        text = "Keep compounding. Freedom is calling."

        # 创建一个临时的 Label 用于居中显示
        welcome_lbl = Label(
            text=text,
            font_size='18sp',
            bold=True,
            color=(1, 1, 1, 1),
            size_hint=(None, None),
            size=(dp(350), dp(60)),
            pos_hint={'center_x': 0.5, 'center_y': 0.5},
            opacity=0
        )

        # 半透明背景
        with welcome_lbl.canvas.before:
            Color(0.1, 0.1, 0.1, 0.85)
            welcome_lbl.bg = RoundedRectangle(pos=welcome_lbl.pos, size=welcome_lbl.size, radius=[dp(15)])

        welcome_lbl.bind(pos=lambda i, v: setattr(i.bg, 'pos', v),
                         size=lambda i, v: setattr(i.bg, 'size', v))

        # 添加到根布局的最上层
        # 注意：kivy app 的 root 是 build() 返回的对象
        self.root_layout.add_widget(welcome_lbl)

        # 动画序列
        anim = Animation(opacity=1, duration=0.8, t='out_quad') + \
               Animation(duration=2.0) + \
               Animation(opacity=0, duration=0.5)

        anim.bind(on_complete=lambda *args: self.root_layout.remove_widget(welcome_lbl))

        # 稍微延迟一点执行，确保界面已经渲染
        Clock.schedule_once(lambda dt: anim.start(welcome_lbl), 0.5)

    def start_auto_refresh_timer(self):
        self.stop_auto_refresh_timer()
        interval = self.backend.config.get('auto_refresh_interval', 60) * 60
        self.refresh_event = Clock.schedule_interval(self.do_auto_refresh, interval)
        print(f"Auto refresh started: every {interval / 60} mins")

    def stop_auto_refresh_timer(self):
        if self.refresh_event:
            self.refresh_event.cancel()
            self.refresh_event = None

    def do_auto_refresh(self, dt):
        print("Auto refreshing...")
        threading.Thread(target=self._background_auto_refresh, daemon=True).start()

    def _background_auto_refresh(self):
        ndx, prices, err = self.backend.get_market_data()
        self.backend.latest_prices = prices
        strategy = self.backend.calculate_strategy(ndx)
        Clock.schedule_once(lambda dt: self._handle_auto_refresh_result(strategy), 0)

    def _handle_auto_refresh_result(self, strategy):
        dash = self.sm.get_screen('dashboard')
        if dash:
            dash.refresh_dashboard(None)

        if strategy['indicator_count'] > 0:
            self.show_alert_popup(strategy)
        else:
            # 常规更新提示
            Toast(text="Market Data Updated", bg_color=(0.15, 0.68, 0.38, 0.9))

    def show_alert_popup(self, strategy):
        content = BoxLayout(orientation='vertical', padding=dp(15), spacing=dp(10))
        content.add_widget(
            Label(text="⚠️ SIGNAL ALERT", font_size='20sp', bold=True, color=(0.91, 0.3, 0.24, 1), size_hint_y=0.2))

        msg = f"Analysis detected {strategy['indicator_count']} buying signals!\n\n"
        msg += "\n".join([f"• {s}" for s in strategy['indicator_details']])
        msg += f"\n\nSuggested Action: Buy ${strategy['total_amount']}"

        content.add_widget(Label(text=msg, font_size='15sp', halign='center'))

        btn = Button(text="Check Now", size_hint_y=0.2, background_color=(0.2, 0.6, 0.86, 1))
        popup = Popup(title='Investment Alert', content=content, size_hint=(0.85, 0.5), auto_dismiss=False)

        def on_check(instance):
            popup.dismiss()
            self.sm.current = 'invest'

        btn.bind(on_press=on_check)
        content.add_widget(btn)
        popup.open()


if __name__ == '__main__':
    InvestmentProApp().run()